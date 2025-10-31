#!/usr/bin/env python3
"""
Export DuckDB table content to an Excel workbook with curated tabs.
"""
import logging
import math
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

import duckdb
import pandas as pd

from logger_config import LoggerConfig

# Tabs that drive ETN CDM filtering logic
TARGET_TABLES = [
    "Allocation",
    "AlternatePart",
    "BillOfMaterial",
    "BOMAlternate",
    "BuyerCode",
    "Calendar",
    "CalendarDate",
    "Constraint",
    "ConstrainAvailable",
    "Customer",
    "DemandOrder",
    "HistoricalDemandHeader",
    "HistoricalDemandActual",
    "HistoricalReceipt",
    "HistoricalReceiptHeader",
    "HistoricalSupplyHeader",
    "HistoricalSupplyActual",
    "IndependentDemand",
    "OnHand",
    "Operation",
    "Part",
    "PartCustomer",
    "PartSource",
    "PartSupplier",
    "PlannerCode",
    "ReferencePart",
    "Site",
    "ScheduledReceipt",
    "Source",
    "SourceConstraint",
    "Supplier",
    "SupplyOrder",
    "Routing",
]

FIELD_CATEGORY_METADATA = [
    (
        "Identifier",
        "Id or part of the Id to uniquely identify a record (ex. Sales Order and Sales Order Line)",
    ),
    (
        "Critical",
        "These elements perform a core function (in a maestro context, missing/incorrect data in these "
        "fields leads to junk/no results) (ex. Purchase Order Confirmation Date)",
    ),
    (
        "Functional Enabler",
        "Element is not needed for core functionality, but if it is missing, some capability(s) will "
        "not function properly (ex. Part ABC/XYZ indicator)",
    ),
    (
        "Optional/Reference",
        "Reference data helps users understand context or makes filtering and grouping easier, but "
        "otherwise does not impact function (ex. Supplier name)",
    ),
]

CRITICAL_NAME_SUBSTRINGS = ("Name", "Site", "Date", "LeadTime", "Number")


class ExcelExporter:
    """Encapsulates the DuckDB → Excel export process."""

    ETN_MATCH_STATUSES = ("ETN_ONLY", "MATCHED")

    def __init__(self, db_path: Path, output_path: Path, overwrite: bool, logger: logging.Logger):
        self.db_path = Path(db_path)
        self.output_path = Path(output_path)
        self.overwrite = overwrite
        self.logger = logger

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #
    def run(self) -> bool:
        """Execute the export pipeline."""
        if not self._validate_paths():
            return False

        payload: Optional[Dict[str, pd.DataFrame]] = None

        try:
            with duckdb.connect(str(self.db_path)) as con:
                self.logger.info("Connected to database: %s", self.db_path)
                payload = self._collect_payload(con)
                self._write_workbook(payload)
        except Exception as exc:  # pragma: no cover - top level guard
            self.logger.error("Export failed: %s", exc)
            return False

        self._log_summary(payload)
        return True

    # ------------------------------------------------------------------ #
    # Data loading
    # ------------------------------------------------------------------ #
    def _collect_payload(self, con: duckdb.DuckDBPyConnection) -> Dict[str, pd.DataFrame]:
        self.logger.info("Exporting tables, columns, and mappings without target table filtering")

        tables_df, table_desc_lookup = self._load_tables(con)
        columns_df, raw_columns_df = self._load_columns(con)

        comments_df = self._load_doc_comments(con)
        cdm_augmentation_df = self._load_cdm_augmentation(con)
        mappings_df = self._load_etn_doc_mappings(con)

        target_tables_upper = [name.upper() for name in TARGET_TABLES]
        key_field_lookup = self._build_key_lookup(raw_columns_df, target_tables_upper)

        etn_cdm_df = self._prepare_etn_cdm(
            con,
            table_desc_lookup=table_desc_lookup,
            key_field_lookup=key_field_lookup,
            target_tables_upper=target_tables_upper,
        )

        field_category_df = pd.DataFrame(FIELD_CATEGORY_METADATA, columns=["Category Name", "Description"])
        summarized_cdm_df = self._load_summarized_cdm(con)

        return {
            "tables": tables_df,
            "columns": columns_df,
            "comments": comments_df,
            "cdm_aug": cdm_augmentation_df,
            "mappings": mappings_df,
            "etn_cdm": etn_cdm_df,
            "field_category": field_category_df,
            "etn_cdm_summary": summarized_cdm_df,
        }

    def _load_tables(
        self,
        con: duckdb.DuckDBPyConnection,
    ) -> Tuple[pd.DataFrame, Dict[str, str]]:
        self.logger.info("Querying tables data...")
        tables_df = con.execute(
            """
            SELECT
                id,
                name AS table_name,
                description,
                calculated_fields_description,
                created_at
            FROM knx_doc_tables
            ORDER BY name
        """
        ).fetchdf()

        self.logger.info("Found %d tables", len(tables_df))
        if tables_df.empty:
            self.logger.warning("No table metadata available to build Maestro description lookup")
            return tables_df, {}

        table_desc_lookup = {
            str(name).strip().upper(): desc
            for name, desc in zip(tables_df["table_name"], tables_df["description"])
            if pd.notna(name)
        }
        self.logger.info("Prepared table description lookup for %d tables", len(table_desc_lookup))
        return tables_df, table_desc_lookup

    def _load_columns(
        self,
        con: duckdb.DuckDBPyConnection,
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        self.logger.info("Querying columns data from knx_doc_extended table...")
        columns_df = con.execute(
            """
            SELECT
                id,
                table_id,
                table_name,
                field_name,
                description,
                data_type,
                is_key,
                is_calculated,
                referenced_table,
                is_extended,
                display_on_export,
                created_at,
                referenced_table_id,
                display_order
            FROM knx_doc_extended
            ORDER BY display_order
        """
        ).fetchdf()

        raw_columns_df = columns_df.copy()

        id_numeric = pd.to_numeric(columns_df["id"], errors="coerce")
        field_names = columns_df["field_name"].astype('string').fillna("").str.lstrip()
        extended_mask = id_numeric.notna() & id_numeric.mod(1).ne(0)
        field_names = field_names.mask(extended_mask, "    " + field_names)
        columns_df["field_name"] = field_names

        self.logger.info("Found %d columns", len(columns_df))
        return columns_df.drop(columns=["display_order"]), raw_columns_df

    def _load_doc_comments(self, con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
        self.logger.info("Querying trl_doc_augmentation data...")
        comments_query = """
            SELECT
                k.table_name,
                k.field_name,
                c.field_sub_domain,
                c.field_view,
                c.field_business_name,
                c.sap_table,
                c.sap_field,
                c.trillium_comments
            FROM trl_doc_augmentation AS c
            INNER JOIN knx_doc_extended AS k
                ON c.table_name = k.table_name
                AND c.field_name = k.field_name
            ORDER BY k.table_name, k.field_name
        """

        try:
            comments_df = con.execute(comments_query).fetchdf()
        except duckdb.CatalogException:
            self.logger.warning("trl_doc_augmentation table not found; creating empty export tab")
            comments_df = pd.DataFrame(
                columns=[
                    "table_name",
                    "field_name",
                    "field_sub_domain",
                    "field_view",
                    "field_business_name",
                    "sap_table",
                    "sap_field",
                    "trillium_comments",
                ]
            )

        self.logger.info("Found %d trl_doc_augmentation records", len(comments_df))
        return comments_df

    def _load_cdm_augmentation(self, con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
        self.logger.info("Querying trl_cdm_augmentation data...")
        try:
            cdm_augmentation_df = con.execute(
                """
                SELECT
                    domain,
                    domain_description,
                    entity,
                    entity_description,
                    applications
                FROM trl_cdm_augmentation
                ORDER BY domain, entity
            """
            ).fetchdf()
        except duckdb.CatalogException:
            self.logger.warning("trl_cdm_augmentation table not found; creating empty export tab")
            cdm_augmentation_df = pd.DataFrame(
                columns=["domain", "domain_description", "entity", "entity_description", "applications"]
            )

        self.logger.info("Found %d trl_cdm_augmentation records", len(cdm_augmentation_df))
        return cdm_augmentation_df

    def _load_etn_doc_mappings(self, con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
        self.logger.info("Querying ETN doc mappings data...")
        mappings_df = con.execute(
            """
            SELECT
                knx_table,
                original_tab,
                source_table,
                source_field,
                special_extract_logic,
                transformation_table_name,
                constant_value,
                target_table,
                target_field,
                example_value,
                notes,
                key,
                show_output,
                sort_output
            FROM etn_doc_mappings
            ORDER BY id
        """
        ).fetchdf()

        self.logger.info("Found %d ETN doc mappings", len(mappings_df))
        return mappings_df

    def _load_summarized_cdm(self, con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
        self.logger.info("Querying summarized ETN CDM data...")
        try:
            df = con.execute(
                """
                SELECT
                    domain,
                    domain_description,
                    entity,
                    entity_description,
                    keys,
                    relationships,
                    applications
                FROM etn_cdm
                ORDER BY domain, entity
            """
            ).fetchdf()
        except duckdb.CatalogException:
            self.logger.warning("etn_cdm table not found; creating empty export tab")
            df = pd.DataFrame(
                columns=[
                    "domain",
                    "domain_description",
                    "entity",
                    "entity_description",
                    "keys",
                    "relationships",
                    "applications",
                ]
            )

        self.logger.info("Found %d summarized ETN CDM rows", len(df))
        if not df.empty and "relationships" in df.columns:
            df["relationships"] = df["relationships"].apply(self._format_relationships_cell)
        return df

    @staticmethod
    def _format_relationships_cell(value: object) -> str:
        if value is None or pd.isna(value):
            return ""

        text = str(value).strip()
        if not text:
            return ""

        if "\n" in text:
            return text

        parts = [part.strip() for part in text.split(",") if part.strip()]
        return "\n".join(parts)

    def _build_key_lookup(
        self,
        raw_columns_df: pd.DataFrame,
        target_tables_upper: Iterable[str],
    ) -> Iterable[Tuple[str, str]]:
        target_tables_upper_set = set(target_tables_upper)
        key_flags = (
            raw_columns_df["is_key"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .isin({"yes", "true", "y", "1"})
        )

        key_columns_df = raw_columns_df[
            raw_columns_df["table_name"].astype('string').fillna("").str.upper().isin(target_tables_upper_set) & key_flags
        ].copy()

        key_columns_df["table_name_upper"] = key_columns_df["table_name"].astype(str).str.upper()
        key_columns_df["field_name_upper"] = key_columns_df["field_name"].astype(str).str.strip().str.upper()

        key_field_lookup = set(zip(key_columns_df["table_name_upper"], key_columns_df["field_name_upper"]))
        self.logger.info("Identified %d target table key fields from knx_doc_extended", len(key_field_lookup))
        return key_field_lookup

    def _prepare_etn_cdm(
        self,
        con: duckdb.DuckDBPyConnection,
        *,
        table_desc_lookup: Dict[str, str],
        key_field_lookup: Iterable[Tuple[str, str]],
        target_tables_upper: Iterable[str],
    ) -> pd.DataFrame:
        self.logger.info("Querying ETN CDM data...")
        self.logger.info(
            "Will filter ETN CDM export to match statuses: %s (Maestro keys retained regardless of status)",
            ", ".join(self.ETN_MATCH_STATUSES),
        )

        etn_cdm_df = con.execute(
            """
            SELECT
                domain_name,
                canonical_entity_name,
                maestro_table_name,
                maestro_table_description,
                canonical_attribute_name,
                maestro_field_name,
                erp_technical_table_name,
                erp_technical_field_name,
                maestro_field_description,
                maestro_data_type,
                maestro_is_key,
                information_only,
                standard_maestro_field,
                add_to_etl,
                default_value,
                example_value,
                erp_tcode,
                erp_screen_name,
                erp_screen_field_name,
                erp_technical_table_name_secondary,
                etl_logic,
                etl_transformation_table,
                notes,
                field_output_order,
                match_status,
                match_tier,
                match_details,
                sap_augmentation_strategy
            FROM etn_cdm_mappings
        """
        ).fetchdf()

        self.logger.info("Found %d ETN CDM records before applying filters", len(etn_cdm_df))

        if not etn_cdm_df.empty and table_desc_lookup and "maestro_table_name" in etn_cdm_df.columns:
            maestro_table_normalized = (
                etn_cdm_df["maestro_table_name"].astype('string').fillna("").str.strip().str.upper()
            )
            existing_desc = etn_cdm_df["maestro_table_description"].astype('string').fillna("").str.strip()
            missing_desc_mask = existing_desc.eq("")
            mapped_descriptions = maestro_table_normalized.map(table_desc_lookup)
            fill_mask = missing_desc_mask & mapped_descriptions.notna()
            if fill_mask.any():
                etn_cdm_df.loc[fill_mask, "maestro_table_description"] = mapped_descriptions[fill_mask]
                self.logger.info(
                    "Filled Maestro table descriptions for %d ETN CDM records using knx_doc_tables metadata",
                    int(fill_mask.sum()),
                )

        key_field_lookup = set(key_field_lookup)
        if not etn_cdm_df.empty and key_field_lookup:
            canonical_upper = etn_cdm_df["canonical_entity_name"].astype('string').fillna("").str.upper()
            maestro_field_upper = etn_cdm_df["maestro_field_name"].astype('string').fillna("").str.strip().str.upper()
            key_mask = pd.Series(
                [(entity, field) in key_field_lookup for entity, field in zip(canonical_upper, maestro_field_upper)],
                index=etn_cdm_df.index,
            )
            key_updates = int(key_mask.sum())
            if key_updates:
                etn_cdm_df.loc[key_mask, "maestro_is_key"] = True
                self.logger.info(
                    "Applied Maestro key flag to %d ETN CDM records based on knx_doc_extended keys",
                    key_updates,
                )

        if not etn_cdm_df.empty and TARGET_TABLES:
            self.logger.info(
                "Filtering ETN CDM export to target tables: %s and match statuses: %s "
                "(retaining Maestro keys regardless of status)",
                ", ".join(TARGET_TABLES),
                ", ".join(self.ETN_MATCH_STATUSES),
            )

            original_count = len(etn_cdm_df)
            canonical_match = etn_cdm_df["canonical_entity_name"].fillna("").str.upper().isin(target_tables_upper)
            status_match = (
                etn_cdm_df["match_status"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
                .isin(self.ETN_MATCH_STATUSES)
            )
            maestro_key_flags = (
                etn_cdm_df["maestro_is_key"]
                .astype('string')
                .fillna("")
                .str.strip()
                .str.upper()
                .isin({"TRUE", "YES", "Y", "1"})
            )

            filter_mask = canonical_match & (status_match | maestro_key_flags)
            etn_cdm_df = etn_cdm_df[filter_mask]

            key_only_retained = int((canonical_match & maestro_key_flags & ~status_match).sum())
            self.logger.info(
                "Filtered ETN CDM records from %d to %d (retained %d Maestro key fields outside status filter)",
                original_count,
                len(etn_cdm_df),
                key_only_retained,
            )

        etn_cdm_df = etn_cdm_df.rename(
            columns={
                "domain_name": "Domain Name",
                "canonical_entity_name": "Canonical Entity Name",
                "maestro_table_name": "Maestro Table Name",
                "maestro_table_description": "Maestro Table Description",
                "canonical_attribute_name": "Canonical Attribute Name",
                "maestro_field_name": "Maestro Field Name",
                "erp_technical_table_name": "ERP Technical Table Name",
                "maestro_field_description": "Maestro Field Description",
                "maestro_data_type": "Maestro Data Type",
                "maestro_is_key": "Maestro Is Key",
                "information_only": "Information Only",
                "standard_maestro_field": "Standard Maestro Field",
                "add_to_etl": "Add to ETL",
                "default_value": "Default Value",
                "example_value": "Example Value",
                "erp_tcode": "ERP TCode",
                "erp_screen_name": "ERP Screen Name",
                "erp_screen_field_name": "ERP Screen Field Name",
                "erp_technical_field_name": "ERP Technical Field Name",
                "erp_technical_table_name_secondary": "ERP Technical Table Name Secondary",
                "etl_logic": "ETL Logic",
                "etl_transformation_table": "ETL Transformation Table",
                "notes": "Notes",
                "field_output_order": "Field Output Order",
                "match_status": "Match Status",
                "match_tier": "Match Tier",
                "match_details": "Match Details",
                "sap_augmentation_strategy": "SAP Augmentation Strategy",
            }
        )

        if "Domain Name" in etn_cdm_df.columns:
            reordered = ["Domain Name"] + [col for col in etn_cdm_df.columns if col != "Domain Name"]
            etn_cdm_df = etn_cdm_df[reordered]

        if "Maestro Is Key" in etn_cdm_df.columns:
            maestro_key_series = (
                etn_cdm_df["Maestro Is Key"].astype('string').fillna("").str.strip().str.lower()
            )
            etn_cdm_df["Maestro Is Key"] = maestro_key_series.isin({"true", "yes", "y", "1", "t"})

        sort_columns = ["Maestro Table Name", "Maestro Is Key", "Maestro Field Name"]
        missing_sort_cols = [col for col in sort_columns if col not in etn_cdm_df.columns]
        if missing_sort_cols:
            self.logger.warning(
                "Unable to apply full ETN CDM sorting; missing columns: %s",
                ", ".join(missing_sort_cols),
            )
        else:
            etn_cdm_df = etn_cdm_df.sort_values(
                by=["Maestro Table Name", "Maestro Is Key", "Maestro Field Name"],
                ascending=[True, False, True],
                kind="mergesort",
            ).reset_index(drop=True)
            self.logger.info("Ordered ETN CDM records by Maestro table name, key flag, then field name")

        if not etn_cdm_df.empty:
            etn_cdm_df["Field Category"] = etn_cdm_df.apply(self._determine_field_category, axis=1)
            self.logger.info("Assigned Field Category values to ETN CDM records")

        return etn_cdm_df

    # ------------------------------------------------------------------ #
    # Workbook generation
    # ------------------------------------------------------------------ #
    def _write_workbook(self, payload: Dict[str, pd.DataFrame]) -> None:
        self.logger.info("Writing to Excel file: %s", self.output_path)
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            self._write_sheet(writer, "knx_doc_tables", payload["tables"])
            self._write_sheet(writer, "knx_doc_extended", payload["columns"])
            self._write_sheet(writer, "etn_doc_mappings", payload["mappings"])
            self._write_sheet(
                writer,
                "ETN CDM Mappings",
                payload["etn_cdm"],
                empty_columns=list(payload["etn_cdm"].columns)
                if not payload["etn_cdm"].empty
                else [
                    "Domain Name",
                    "Canonical Entity Name",
                    "Maestro Table Name",
                    "Maestro Table Description",
                    "Canonical Attribute Name",
                    "Maestro Field Name",
                ],
            )
            self._write_sheet(
                writer,
                "ETN CDM",
                payload["etn_cdm_summary"],
                empty_columns=list(payload["etn_cdm_summary"].columns)
                if not payload["etn_cdm_summary"].empty
                else [
                    "domain",
                    "domain_description",
                    "entity",
                    "entity_description",
                    "keys",
                    "relationships",
                    "applications",
                ],
            )
            self._write_sheet(writer, "Field Category", payload["field_category"])
            self._write_sheet(writer, "trl_doc_augmentation", payload["comments"])
            self._write_sheet(writer, "trl_cdm_augmentation", payload["cdm_aug"])

            self._format_workbook(writer)

    def _write_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        *,
        empty_columns: Optional[Iterable[str]] = None,
    ) -> None:
        if df.empty and empty_columns:
            dataframe = pd.DataFrame(columns=list(empty_columns))
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            self.logger.info("Created empty '%s' tab (no data available)", sheet_name)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self.logger.info("Written %s rows to '%s' tab", len(df), sheet_name)

    def _format_workbook(self, writer: pd.ExcelWriter) -> None:
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        def estimate_lines(cell_text: Optional[str], column_width: Optional[float]) -> int:
            if not cell_text:
                return 1
            approx_chars = max(int((column_width or 10) * 0.9), 10)
            total_lines = 0
            for raw_line in cell_text.splitlines() or [""]:
                stripped = raw_line.strip()
                if not stripped:
                    total_lines += 1
                    continue
                total_lines += max(1, math.ceil(len(stripped) / approx_chars))
            return max(total_lines, cell_text.count("\n") + 1)

        for sheet_name, worksheet in writer.sheets.items():
            # Configure column widths and wrapping
            for column in worksheet.columns:
                cells = list(column)
                if not cells:
                    continue

                column_letter = cells[0].column_letter
                header_value = cells[0].value
                header_lower = str(header_value).strip().lower() if header_value else ""
                is_description_column = "description" in header_lower
                is_field_name_column = header_lower == "field_name"

                max_length = 0
                for cell in cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    value_length = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, value_length)

                if sheet_name == "knx_doc_extended" and is_field_name_column:
                    worksheet.column_dimensions[column_letter].width = 60
                elif is_description_column:
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 80)
                else:
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)

            # Adjust row heights
            try:
                header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                header_row = []

            description_columns = {
                cell.column_letter
                for cell in header_row
                if cell.value and "description" in str(cell.value).strip().lower()
            }

            for row in worksheet.iter_rows():
                max_lines_in_row = 1
                row_number = row[0].row if row else 0
                for cell in row:
                    if cell.value is None:
                        continue
                    column_width = worksheet.column_dimensions[cell.column_letter].width or 10
                    lines_needed = estimate_lines(str(cell.value), column_width)
                    if sheet_name == "knx_doc_extended" and cell.column_letter in description_columns:
                        lines_needed = max(lines_needed, 2)
                    max_lines_in_row = max(max_lines_in_row, lines_needed)
                worksheet.row_dimensions[row_number].height = max(20, max_lines_in_row * 15)

            # Auto-filter
            if worksheet.max_row > 1:
                max_col_letter = get_column_letter(worksheet.max_column)
                worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"

            # Hide ID-centric columns on the extended sheet
            if sheet_name == "knx_doc_extended":
                hidden_headers = {"id", "table_id", "referenced_table_id", "display_on_export"}
                for col_idx, column in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
                    header = column[0].value
                    if header and str(header).lower() in hidden_headers:
                        worksheet.column_dimensions[get_column_letter(col_idx)].hidden = True

            worksheet.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #
    def _determine_field_category(self, row: pd.Series) -> Optional[str]:
        maestro_is_key = bool(row.get("Maestro Is Key", False))
        if maestro_is_key:
            return "Identifier"

        match_status = (row.get("Match Status") or "").strip().upper()
        if match_status == "ETN_ONLY":
            return "Optional/Reference"

        maestro_field_name = row.get("Maestro Field Name") or ""
        if any(substring in maestro_field_name for substring in CRITICAL_NAME_SUBSTRINGS) or "LT" in maestro_field_name:
            return "Critical"

        if match_status == "MATCHED":
            return "Functional Enabler"

        return None

    def _validate_paths(self) -> bool:
        if not self.db_path.exists():
            self.logger.error("Database file %s not found", self.db_path)
            return False

        if self.output_path.exists():
            if not self.overwrite:
                self.logger.error(
                    "Output file %s already exists. Use overwrite=True to replace it.",
                    self.output_path,
                )
                return False
            try:
                self.logger.info("Output file %s exists, overwriting...", self.output_path)
                self.output_path.unlink()
            except Exception as exc:  # pragma: no cover - defensive
                self.logger.error("Failed to delete existing file %s: %s", self.output_path, exc)
                return False

        return True

    def _log_summary(self, payload: Optional[Dict[str, pd.DataFrame]]) -> None:
        if not payload:
            return

        self.logger.info("Excel export completed successfully!")
        self.logger.info("Output file: %s", self.output_path)
        self.logger.info("Tables exported: %d", len(payload["tables"]))
        self.logger.info("Columns exported: %d", len(payload["columns"]))
        self.logger.info("ETN mappings exported: %d", len(payload["mappings"]))


# ---------------------------------------------------------------------- #
# Module API
# ---------------------------------------------------------------------- #
def export_to_excel(db_path: str = "mappings.duckdb", output_file: str = "tables_export.xlsx", overwrite: bool = False):
    """Entry point used by CLI and tests."""
    logger_config = LoggerConfig(name="ExcelExporter", log_level=logging.INFO, log_file="export.log")
    logger = logger_config.get_logger()
    logger.info("Starting Excel export process")

    exporter = ExcelExporter(db_path=Path(db_path), output_path=Path(output_file), overwrite=overwrite, logger=logger)
    return exporter.run()


def main():
    """Main function to run the export."""
    print("DuckDB to Excel Exporter")
    print("=" * 40)

    output_file = "kinaxis_tables_export.xlsx"
    success = export_to_excel(output_file=output_file, overwrite=True)

    if success:
        print(f"✅ Export successful! File saved as: {output_file}")
    else:
        print("❌ Export failed. Check the logs for details.")


if __name__ == "__main__":
    main()
